#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
    NAME
        pandas_practice.py

    DESCRIPTION
        to practice pandas
        # ref: [dataguru.cn]python05.pdf
        # ref: pandas data structure and data operation
                https://blog.csdn.net/pipisorry/article/details/18010307

    MODIFIED  (MM/DD/YY)
        Na  11/20/2018

"""
__VERSION__ = "1.0.0.11202018"


# imports
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os, os.path
import pprint as pp

# consts
CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(CURRENT_PATH, r'data\week5_data_after_housework')
# CURRENT_PATH, DATA PATH here is for running code snippet in Python Console by ^+Alt+E
# CURRENT_PATH = os.getcwd()
# DATA_PATH = os.path.join(CURRENT_PATH, r'pda\data\week5_data_after_housework')

# functions
def main():
    # Series in pandas
    obj = pd.Series([4, 7, -5, 3])
    print('pd.Series obj:\n{}\nobj.values: {}\nobj.index : {}'.format(
        obj, obj.values, obj.index ))
    print('obj.describe:\n{}'.format(obj.describe()))

    obj2 = pd.Series([4, 7, -5, 3], index=['d', 'b', 'a', 'c'])
    print('obj2:\n{}\nobj2.index: {}'.format(obj2, obj2.index))
    print("obj2['a']: {}".format(obj2['a']))
    obj2['d'] = 6
    print('obj2:\n{}\nobj2.index: {}'.format(obj2, obj2.index))
    print("obj2[['c', 'a', 'd']]:\n{}".format(obj2[['c', 'a', 'd']]))
    print('\nobj2 > 0:\n{}'.format(obj2[obj2 > 0]))
    print('\nobj2 * 2:\n{}]'.format(obj2 * 2))
    print('np.exp(obj2):\n{}'.format(np.exp(obj2)))
    print("'b' in obj2? {}".format('b' in obj2))
    print("'e' in obj2? {}".format('e' in obj2))

    sdata = {'Ohio': 35000, 'Texas': 71000, 'Oregon': 16000, 'Utah': 5000}
    obj3 = pd.Series(sdata)
    print('\nobj3:\n{}'.format(obj3))

    states = ['California', 'Ohio', 'Oregon', 'Texas']
    obj4 = pd.Series(sdata, index=states)
    print('\nobj4:\n{}'.format(obj4))
    print('pd.isnull(obj4):\n{}'.format(pd.isnull(obj4)))
    print('\nobj4.isnull():\n{}'.format(obj4.isnull()))
    print('pd.isnull(obj4) == obj4.isnull()?: {}'.format(np.array_equal(pd.isnull(obj4), obj4.isnull())))
    print('\npd.notnull(obj4):\n{}'.format(pd.notnull(obj4)))

    print('\nobj3:\n{}\nobj4:\n{}\nobj3+obj4:\n{}'.format(obj3, obj4, obj3+obj4))

    obj4.name = 'population'
    obj4.index.name = 'state'
    print('\nobj4:\n{}'.format(obj4))

    print('\nobj before setting index explicitly:\n{}'.format(obj))
    obj.index = ['Bob', 'Steve', 'Jeff', 'Ryan']
    print('\nobj after setting index explicitly:\n{}'.format(obj))

    # dataframe
    data = {'state': ['Ohio', 'Ohio', 'Ohio', 'Nevada', 'Nevada'],
            'year': [2000, 2001, 2002, 2001, 2002],
            'pop': [1.5, 1.7, 3.6, 2.4, 2.9]}
    frame = pd.DataFrame(data)
    print('\nframe:\n{}'.format(frame))
    print('frame after designated columns:\n{}'.format(pd.DataFrame(data, columns=['year', 'state', 'pop'])))

    frame2 = pd.DataFrame(data=data, index=['one', 'two', 'three', 'four', 'five'],
                          columns=['year', 'state', 'pop', 'debt'])
    # get column, row
    print('frame2 after designated index and columns:\n{}\nframe2.columns:\t{}'.format(frame2, frame2.columns))
    print("frame2['state']:\n{}".format(frame2['state']))
    print('frame2.year:\n{}'.format(frame2.year))
    print('frame2.ix["three"]:\ttype of it: {}\n{}'.format(type(frame2.ix['three']), frame2.ix['three']))

    # setting values row-based, columned, update values
    frame2['debt'] = 16.5
    print('frame2 after assigned debt:\n{}'.format(frame2))
    frame2['debt'] = np.arange(5.)
    print('frame2 after assigned debt:\n{}'.format(frame2))

    frame2['pop']['three'] = 8.0
    print("Data updated after frame2['pop']['three'] = 8.0:\n{}".format(frame2))
    frame2.ix['three']['pop'] = 9.0
    print("Data NOT changed after frame2.ix['three']['pop'] = 9.0:\n{}".format(frame2))
    print('Following data setting methods works:')
    frame2.iat[2, 2] = 9.0
    frame2.loc['three', 'pop'] = 10.0
    frame2.at['three', 'pop'] = 20.0

    va1 = pd.Series([-1.2, -1.5, -1.7], index=['two', 'four', 'five'])
    frame2['debt'] = va1
    print('frame2 after assigned debt partially:\n{}'.format(frame2))

    # add columns, del columns
    frame2['eastern'] = frame2.state == 'Ohio'
    print('frame2 which has eastern?\n{}'.format(frame2))
    del frame2['eastern']
    print('frame2 columns\n{}'.format(frame2.columns))

    pop = {'Nevada': {2001: 2.4, 2002: 2.9},
           'Ohio': {2000: 1.5, 2001: 1.7, 2002: 3.6}}
    frame3 = pd.DataFrame(pop)
    print('frame3:\n{}\nframe3.T:\n{}'.format(frame3, frame3.T))

    # add index.name, columns.name
    frame3.index.name = 'year'
    frame3.columns.name = 'state'
    print('frame3:\n{}\ntype of frame3.values:\t{}\nframe3.values:\n{}'.format(frame3, type(frame3.values), frame3.values))
    print('frame2.values:\n{}'.format(frame2.values))

    # drop data on an axis
    # for Series
    frame2['debt'].drop('one')
    frame2.drop('debt', axis=1)
    frame2.drop(['state', 'pop'], axis=1)
    frame2.drop(['one', 'two'], axis=0)

    # add dfs
    frame + frame2
    (frame + frame2).fillna(0)
    # # error: TypeError: cannot concatenate 'str' and 'float' objects
    # frame.add(frame2, fill_value=0)
    # works:
    pd.concat([frame.iloc[range(5)], frame2[['year', 'pop']]])


    # index object
    obj = pd.Series(range(3), index=['b', 'a', 'c'])
    index = obj.index
    print('obj:\n{}\nobj.index:\n{}'.format(obj, obj.index))
    print('index[1:]: {}'.format(index[1:]))
    # index is immutable
    # index[1] = 'd'
    # print('index after set index[1] = "d": {}'.format(index))

    index2 = pd.Index(np.arange(3))
    print('index2:\n{}'.format(index2))
    obj2 = pd.Series([-1.5, -2.5, 0], index=index)
    print('obj2:\n{}'.format(obj2))
    print('is obj2.index is index? {}'.format(obj2.index is index))
    # index must have same length as data
    # obj2.index = index2[1:]

    # checking existence
    print("'Ohio' in frame3.columns? {}".format('Ohio' in frame3.columns))
    print('2003 in frame3.index? {}'.format(2003 in frame3.index))

    # reading data
    # text
    f1 = os.path.join(DATA_PATH, 'ex1.csv')
    df = pd.read_csv(f1)
    print('df in {}:\n{}'.format(f1, df))
    df11 = pd.read_table(f1, sep=',')
    print('df in {} pd.read_table:\n{}'.format(f1, df11))
    f2 = os.path.join(DATA_PATH, 'ex2.csv')
    print('reading {} ...'.format(f2))
    print('\tpd.read_csv with header=None:\n{}'.format(pd.read_csv(f2, header=None)))
    print('\tpd.read_csv with neames:\n{}'.format(pd.read_csv(f2, names=['a', 'b', 'c', 'd', 'message'])))
    names = ['a', 'b', 'c', 'd', 'message']
    print('\tpd.read_csv with names and index_col:\n{}'.format(pd.read_csv(f2, names=names, index_col='message')))

    f_mindex = os.path.join(DATA_PATH, 'csv_mindex.csv')
    parsed = pd.read_csv(f_mindex, index_col=['key1', 'key2'])
    print('pd.read_csv with multiple indexes:\n{}'.format(parsed))

    f3 = os.path.join(DATA_PATH, 'ex3.txt')
    print('list(open(f3)):\n')
    import pprint as pp
    pp.pprint(list(open(f3)))
    result = pd.read_csv(f3, sep='\s+')
    print('same file read by pd.read_csv with sep="\s+":\n{}'.format(result))

    f4 = os.path.join(DATA_PATH, 'ex4.csv')
    print('pd.read_csv(f4, skiprows=[0, 2, 3]):\n{}'.format(pd.read_csv(f4, skiprows=[0, 2, 3])))

    f5 = os.path.join(DATA_PATH, 'ex5.csv')
    result = pd.read_csv(f5)
    print('f5:\n{}'.format(result))
    sentinels = {'message': ['foo', 'NaN'], 'something': ['two']}
    result55 = pd.read_csv(f5, na_values=sentinels)
    print('f5:\n{}'.format(result55))

    # read data line by line
    f6 = os.path.join(DATA_PATH, 'ex6.csv')
    result = pd.read_csv(f6)
    print('contents in {}:\n{}'.format(f6, result))
    result_5rows = pd.read_csv(f6, nrows=5)
    print('first 5 rows in {}:\n{}'.format(f6, result_5rows))
    # chunksize: each chunk has 1000 rows,
    #   the content of the file is divided into: chunksize * number_of_chunks.
    chunker = pd.read_csv(f6, chunksize=1000)
    print('chunker with pd.read_csv(f6, chunksize=1000):\n{}'.format(chunker))

    tot = pd.Series([])
    for piece in chunker:
        tot = tot.add(piece['key'].value_counts(), fill_value=0)
        # print("for each piece in chunker:\ttot = tot.add(piece['key'].value_counts(), fill_value=0):\n{}".format(tot))
    tot = tot.sort_values(ascending=False)
    print('total after order(ascending=False):\n{}'.format(tot))
    print('First 10 line of tot:\n{}'.format(tot[:10]))

    # write to file
    data = pd.read_csv(f5)
    data.to_csv(os.path.join(DATA_PATH, 'out.csv'))
    data.to_csv(sys.stdout, sep='|')
    data.to_csv(sys.stdout, na_rep='NULL')
    data.to_csv(sys.stdout, index=False, header=False)
    data.to_csv(sys.stdout, index=False, columns=['a', 'b', 'c'])

    # write dates into file
    dates = pd.date_range('1/1/2000', periods=7)
    ts = pd.Series(np.arange(7), index=dates)
    print('Series:\n{}\n\tdates as index:\n{}'.format(ts, dates))
    f_tseries = os.path.join(DATA_PATH, 'tseries.csv')
    ts.to_csv(f_tseries)

    print('Read dates series from file:\n{}'.format(pd.Series.from_csv(f_tseries)))

    # process seperators manually
    f7 = os.path.join(DATA_PATH, 'ex7.csv')
    import csv
    with open(f7) as f:
        reader = csv.reader(f)
        print('type(reader): {}'.format(type(reader)))
        for line in reader:
            print('line from csv.reader(f7): {}'.format(line))
        f.seek(0)
        lines = list(csv.reader(f))
        print('lines in another csv.reader(f): {}'.format(lines))
        header, values = lines[0], lines[1:]
        data_dict = {h: v for h, v in zip(header, zip(*values))}

    class my_dialect(csv.Dialect):
        lineterminator, delimiter = '\n', ';'
        quotechar, quoting = '"', csv.QUOTE_MINIMAL

    fmydata = os.path.join(DATA_PATH, 'mydata.csv')
    with open(fmydata, 'w') as f:
        writer = csv.writer(f, dialect=my_dialect)
        writer.writerow(('one', 'two', 'three'))
        writer.writerow(('1', '2', '3'))
        writer.writerow(('4', '5', '6'))
        writer.writerow(('7', '8', '9'))
    print('pd.read_table(fmydata, sep=";"):\n{}'.format(pd.read_table(fmydata, sep=";")))

    # Excel data
    # write to Excel Workbook: Sheet
    import xlrd, xlwt
    fexcel = os.path.join(DATA_PATH, 'workbook.xls')
    wb = xlwt.Workbook()
    print('xlwt.Workbook():\t{}'.format(wb))
    wb.add_sheet('first_sheet', cell_overwrite_ok=True)
    print('wb.get_active_sheet():\t{}'.format(wb.get_active_sheet()))

    ws_1 = wb.get_sheet(0)
    print('wb.get_sheet(0):\t{}'.format(ws_1))
    ws_2 = wb.add_sheet('second_sheet')
    print("wb.add_sheet('second_sheet'):\t{}".format(ws_2))

    data = np.arange(1, 65).reshape((8, 8))
    print('Generated data:\n{}'.format(data))

    ws_1.write(0, 0, 100)
    for c in range(data.shape[0]):
        for r in range(data.shape[1]):
            ws_1.write(r, c, data[c, r])
            ws_2.write(r, c, data[r, c])
    wb.save(fexcel)

    # read from Excel Workbook: Sheet
    book = xlrd.open_workbook(fexcel)
    print('xlrd.open_workbook({}):\t{}'.format(fexcel, book))
    print('book.sheet_names():\t{}'.format(book.sheet_names()))
    sh_1 = book.sheet_by_name('first_sheet')
    sh_2 = book.sheet_by_index(1)
    print("book.sheet_by_name('first_sheet'):\t{}".format(sh_1))
    print("book.sheet_by_index(1).name:\t{}".format(sh_2.name))

    print('sh_1.nrows:\t{}\tsh_1.ncols:\t{}'.format(sh_1.nrows, sh_1.ncols))

    c1 = sh_1.cell(0, 0)
    print('value of sh_1.cell(0, 0):\t{}, type:\t{}'.format(c1.value, c1.ctype))
    print('sh_2.row(3):\n\t{}'.format(sh_2.row(3)))
    print('sh_2.col(3):\n\t{}'.format(sh_2.col(3)))
    print('sh_1.col_values():\n\t{}'.format(sh_1.col_values(3, start_rowx=3, end_rowx=7)))
    print('sh_1.row_values():\n\t{}'.format(sh_1.row_values(3, start_colx=3, end_colx=7)))

    for c in xrange(sh_1.ncols):
        for r in xrange(sh_1.nrows):
            print('sh_1.cell({}, {}).value:\t{}'.format(r, c, sh_1.cell(r, c).value))

    # read via pandas: works for both .xls, .xlsx
    xls_file = pd.ExcelFile(fexcel)
    table = xls_file.parse('first_sheet')
    print("table from pd.ExcelFile({}).parse('first_shee'):\n{}".format(fexcel, table))
    table2 = xls_file.parse(1)
    print("table from pd.ExcelFileExcelFile({}).parse(2):\n{}".format(fexcel, table2))

    fexcel2 = os.path.join(DATA_PATH, 'workbook.xlsx')
    xls_file2 = pd.ExcelFile(fexcel2)
    xls_file2.sheet_names
    xls_file2.parse(1)
    xls_file2.parse('first_sheet')

    fopyxl = os.path.join(DATA_PATH, 'workbook.xlsx')
    import openpyxl
    owb = openpyxl.load_workbook(fopyxl)
    ows_1 = owb['first_sheet']
    print('openpyxl.load_workbook().get_sheet_by_name:\t{}'.format(ows_1))

    print('first_sheet: {} rows {} columns'.format(ows_1.max_row, ows_1.max_column))
    for r in ows_1.rows:
        print('ows_1.rows:\t{}'.format([c.value for c in r if c]))
    for v in ows_1.values:
        print('ows_1.values:\t{}'.format(v))


    # JSON data
    obj = """
    {"name": "Wes",
     "places_lived": ["United States", "Spain", "Germany"],
     "pet": null,
     "siblings": [{"name": "Scott", "age": 25, "pet": "Zuko"},
                  {"name": "Katie", "age": 33, "pet": "Cisco"}]
    }
    """
    import json
    result = json.loads(obj)
    print('original data in python:\n{}\njson.loads(obj):\n{}'.format(obj, result))

    asjson = json.dumps(result)
    print('\njson.dumps(result):\n{}'.format(asjson))

    siblings = pd.DataFrame(result['siblings'], columns=['name', 'age'])
    print("pd.DataFrame(result['siblings'...]:\n{}".format(siblings))

    fjson = os.path.join(DATA_PATH, 'out.json')
    with open(fjson, 'w') as f:
        json.dump(result, f)

    # Error: ValueError: arrays must all be same length
    pd.read_json(fjson)
    df_labels = pd.read_json(fhtml)

    # works
    siblings.to_json(fjson)
    pd.read_json(fjson)


    # binary data
    frame = pd.read_csv(f1)
    print('pd.read_csv(f1):\n{}'.format(frame))
    fpickle = os.path.join(DATA_PATH, 'frame_pickle')
    frame.to_pickle(fpickle)
    print('frame is to_pickle() to {}'.format(fpickle))
    with open(fpickle) as f:
        print('Reading content of frame pickled as text:\n{}'.format(f.read()))
    framed_pk = pd.read_pickle(fpickle)
    print('Reading content of frame picked as pd.read_pickle:\n{}'.format(framed_pk))

    # HDF5 data
    fhdf5 = os.path.join(DATA_PATH, 'mydata.h5')
    store = pd.HDFStore(fhdf5)
    store['obj1'] = frame
    store['obj1_col'] = frame['a']
    print('store after pd.HDFStore(fhdf5):\n{}'.format(store))
    print("\nstore['obj1']:\n{}".format(store['obj1']))
    store.close()
    # os.remove(fhdf5)

    # HTML and requests
    import requests
    url = 'https://api.github.com/repos/pydata/pandas/milestones/28/labels'
    resp = requests.get(url)
    print('resp from requests({}):\n{}'.format(url, resp))
    data = json.loads(resp.text)    # data is <type 'list'>
    print('json.loads(resp.text):\n{}'.format(data))

    issue_labels = pd.DataFrame(data)
    # same as
    # issue_labels = pd.DataFrame.from_dict(data)
    print('issue_labels:\n{}'.format(issue_labels))

    # pip install lxml
    # pip install html5lib
    # pip install BeautifulSoup4
    # !! restart PyCharm !!
    # <=> pd.read_html(url) works for good-formatted url.
    # ref: https://stackoverflow.com/questions/34555135/pandas-read-html
    # Not work for https://api.github.com/repos/pydata/pandas/milestones/28/labels:
    #   ValueError: No tables found
    # df_labels = pd.read_html(url)

    # using DB
    import sqlite3
    query = """
    CREATE TABLE test
    (a VARCHAR(20), b VARCHAR(20),
     c REAL,        d INTEGER
     )
    """
    conn = sqlite3.connect(':memory:')
    conn.execute(query)
    conn.commit()

    data = [('Atlanta', 'Georgia', 1.25, 6),
        ('Tallahassee', 'Florida', 2.6, 3),
        ('Sacramento', 'California', 1.7, 5)]
    stmt = 'INSERT INTO test VALUES(?, ?, ?, ?)'

    conn.executemany(stmt, data)
    conn.commit()

    cursor = conn.execute('select * from test')
    rows = cursor.fetchall()
    print('rows read from sqlite3 :memory: DB test table:\n{}'.format(rows))
    print('cursor.description:\n{}'.format(cursor.description))

    df_sql = pd.DataFrame(rows, columns=zip(*cursor.description)[0])
    print('df_sql after pd.DataFrame():\n{}'.format(df_sql))

    import pandas.io.sql as sql
    pd_sql = sql.read_sql('select * from test', conn)
    print('pd_sql after sql.read_sql():\n{}'.format(pd_sql))

    print('df_sql.equals(pd_sql): {}'.format(df_sql.equals(pd_sql)))
    conn.close()




# classes


# main entry
if __name__ == "__main__":
    main()
    